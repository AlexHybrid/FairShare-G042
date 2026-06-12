from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
import os
import io
import sqlite3
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from database import init_db, get_db_connection

# Excel and PDF Generation Imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter

app = Flask(__name__)
app.secret_key = 'your_very_secret_key_here'

UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads'))
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Initialize database on startup
init_db()

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/signup', methods=['POST'])
def signup():
    name = request.form['name']
    email = request.form['email']
    password = request.form['password']
    confirm_password = request.form['confirm_password']
    
    if password != confirm_password:
        flash('Passwords do not match!', 'error')
        return redirect(url_for('index'))
    
    hashed_password = generate_password_hash(password)
    
    conn = get_db_connection()
    try:
        conn.execute('INSERT INTO users (name, email, password) VALUES (?, ?, ?)',
                     (name, email, hashed_password))
        conn.commit()
        flash('Account created successfully! Please log in.', 'success')
    except sqlite3.IntegrityError:
        flash('An account with this email already exists.', 'error')
    finally:
        conn.close()
        
    return redirect(url_for('index'))

@app.route('/login', methods=['POST'])
def login():
    email = request.form['email']
    password = request.form['password']
    
    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
    conn.close()
    
    if user and check_password_hash(user['password'], password):
        session['user_id'] = user['id']
        session['user_name'] = user['name']
        return redirect(url_for('dashboard'))
    else:
        flash('Invalid email or password. Please try again.', 'error')
        return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash('Please log in to access the dashboard.', 'error')
        return redirect(url_for('index'))
    return render_template('dashboard.html', name=session['user_name'])

@app.route('/export')
def export_page():
    if 'user_id' not in session:
        flash('Please log in to access the export page.', 'error')
        return redirect(url_for('index'))
    return render_template('ExportPage.html', name=session['user_name'])

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'success')
    return redirect(url_for('index'))

# --- API ENDPOINTS FOR EXPENSES ---

@app.route('/api/expenses', methods=['GET'])
def get_expenses():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM expenses')
    rows = cursor.fetchall()
    conn.close()

    expenses = []
    for row in rows:
        expenses.append({
            'id':             row['id'],
            'date':           row['date'],
            'room':           row['room'],
            'type':           row['type'],
            'amount':         row['amount'],
            'payment_method': row['payment_method'] or 'Cash',
            'category':       row['category']       or 'Utilities',
            'notes':          row['notes']           or '',
            'receipt_path':   row['receipt_path']    or '',
            'is_recurring':   bool(row['is_recurring']),
            'status':         row['status']          or 'Pending',
        })
    return jsonify(expenses)

@app.route('/api/expenses', methods=['POST'])
def add_expense():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.json
    date           = data.get('date')
    room           = data.get('room')
    type_          = data.get('type')
    amount         = data.get('amount')
    payment_method = data.get('payment_method', 'Cash')
    category       = data.get('category', 'Utilities')
    notes          = data.get('notes', '')
    receipt_path   = data.get('receipt_path', '')
    is_recurring   = 1 if data.get('is_recurring') else 0
    status         = data.get('status', 'Pending')

    if not all([date, room, type_, amount]):
        return jsonify({'error': 'Missing required fields'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO expenses
            (date, room, type, amount, payment_method, category, notes, receipt_path, is_recurring, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (date, room, type_, float(amount), payment_method, category, notes, receipt_path, is_recurring, status))
    conn.commit()
    conn.close()

    return jsonify({'status': 'success'}), 201

@app.route('/api/expenses/upload', methods=['POST'])
def upload_receipt():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    if 'receipt' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['receipt']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_DIR, filename)
        file.save(filepath)
        return jsonify({'receipt_path': filename}), 200
    return jsonify({'error': 'File type not allowed'}), 400

@app.route('/api/expenses/export/excel', methods=['GET'])
def export_excel():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM expenses ORDER BY id ASC')
        rows = cursor.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses Report"

        headers = ["ID", "Date", "Room/Housemate", "Expense Type", "Amount (RM)", "Payment Method", "Category", "Status", "Notes"]
        ws.append(headers)

        header_fill = PatternFill(start_color="111833", end_color="111833", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        for row in rows:
            ws.append([
                row['id'],
                row['date'],
                row['room'],
                row['type'],
                row['amount'],
                row['payment_method'] or 'Cash',
                row['category'] or 'Utilities',
                row['status'] or 'Pending',
                row['notes'] or ''
            ])

        # Formatting data cells
        for row_idx in range(2, len(rows) + 2):
            cell_amount = ws.cell(row=row_idx, column=5)
            cell_amount.number_format = '"RM"#,##0.00'
            cell_amount.alignment = Alignment(horizontal="right")
            
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border
                c.font = Font(name="Segoe UI", size=10)
                if col_idx in [1, 2, 8]:
                    c.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="FairShare_Expenses_Report.xlsx"
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/expenses/export/pdf', methods=['GET'])
def export_pdf():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT date, room, type, category, status, amount FROM expenses ORDER BY id ASC')
        rows = cursor.fetchall()
        conn.close()

        total_amount = sum(row['amount'] for row in rows)

        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )

        styles = getSampleStyleSheet()
        
        style_normal = ParagraphStyle(
            name='NormalStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=11,
            textColor=colors.HexColor('#1e293b')
        )
        style_header = ParagraphStyle(
            name='HeaderStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            leading=12,
            textColor=colors.white,
            alignment=1
        )
        style_right = ParagraphStyle(
            name='RightStyle',
            parent=style_normal,
            alignment=2
        )
        style_center = ParagraphStyle(
            name='CenterStyle',
            parent=style_normal,
            alignment=1
        )

        data = [
            [
                Paragraph("Date", style_header),
                Paragraph("Room", style_header),
                Paragraph("Expense Type", style_header),
                Paragraph("Category", style_header),
                Paragraph("Status", style_header),
                Paragraph("Amount", style_header)
            ]
        ]

        for row in rows:
            data.append([
                Paragraph(row['date'], style_center),
                Paragraph(row['room'], style_center),
                Paragraph(row['type'], style_normal),
                Paragraph(row['category'] or 'Utilities', style_center),
                Paragraph(row['status'] or 'Pending', style_center),
                Paragraph(f"RM{row['amount']:.2f}", style_right)
            ])

        data.append([
            Paragraph("<b>TOTAL</b>", style_right),
            "", "", "", "",
            Paragraph(f"<b>RM{total_amount:.2f}</b>", style_right)
        ])

        t = Table(data, colWidths=[75, 65, 142, 90, 75, 85])
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111833')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#1e293b')),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            ('SPAN', (0, -1), (4, -1)),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f5f9')),
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#cbd5e1')),
            ('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#cbd5e1')),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ])
        t.setStyle(t_style)

        story = []
        style_title = ParagraphStyle(
            name='TitleStyle',
            fontName='Helvetica-Bold',
            fontSize=24,
            leading=28,
            textColor=colors.HexColor('#111833'),
            spaceAfter=6
        )
        style_subtitle = ParagraphStyle(
            name='SubTitleStyle',
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=20
        )

        story.append(Paragraph("FairShare — Expenses Report", style_title))
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        story.append(Paragraph(f"Generated on {gen_time} | Active House Rentals & Bill Splitter", style_subtitle))
        story.append(Spacer(1, 10))
        story.append(t)

        doc.build(story)
        pdf_buffer.seek(0)

        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name="FairShare_Expenses_Report.pdf"
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
